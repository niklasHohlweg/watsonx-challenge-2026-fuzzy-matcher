"""
exporter — build and write a formatted multi-sheet Excel workbook.

Public API:
    build_workbook(results, df_a, extra_cols_a, df_b, extra_cols_b,
                   output_path, active_tiers=None)

Output sheets:
    1. Summary   – tier counts, percentages, metadata
    2. Matched   – rows whose tier is in active_tiers, formatted as an Excel
                   Table with AutoFilter pre-set to show only those tiers.
                   Defaults to Exact only.
    3. Unmatched – rows from File A that received No Match, also as a Table.
"""
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from models.match_result import (
    MatchResult,
    TIER_EXACT,
    TIER_GOOD,
    TIER_POSSIBLE,
    TIER_NO_MATCH,
    TIER_FILL_COLOURS,
    TIERS,
)

# Styling constants
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
_BODY_FONT = Font(name="Calibri", size=10)
_BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

_TIER_FILLS: dict[str, PatternFill] = {
    tier: PatternFill(fill_type="solid", fgColor="FF" + colour)
    for tier, colour in TIER_FILL_COLOURS.items()
}

_MAX_COL_WIDTH = 50
_MIN_COL_WIDTH = 8


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def deduplicate_results(results: list[MatchResult]) -> list[MatchResult]:
    """
    Return a copy of *results* with File B duplicates resolved.

    When two or more File A rows claim the same File B index_b as their
    best match, only the one with the highest combined_score keeps that
    match.  All others are downgraded to No Match (index_b=None, tier
    set to TIER_NO_MATCH) so they appear in the Unmatched sheet instead.

    Rows that already have index_b=None (No Match) are never affected.
    The original list is not mutated.
    """
    import copy

    results = [copy.copy(r) for r in results]  # shallow copy each result

    # Find the best-scoring result for every File B index
    best_for_b: dict[int, MatchResult] = {}
    for r in results:
        if r.index_b is None:
            continue
        existing = best_for_b.get(r.index_b)
        if existing is None or r.combined_score > existing.combined_score:
            best_for_b[r.index_b] = r

    # Downgrade losers
    for r in results:
        if r.index_b is None:
            continue
        if best_for_b.get(r.index_b) is not r:
            r.tier = TIER_NO_MATCH
            r.index_b = None

    return results


def build_workbook(
    results: list[MatchResult],
    df_a: pd.DataFrame,
    extra_cols_a: list[str],
    df_b: pd.DataFrame,
    extra_cols_b: list[str],
    output_path: str,
    active_tiers: list[str] | None = None,
    deduplicate: bool = False,
) -> None:
    """
    Write a formatted Excel workbook to *output_path*.

    Parameters
    ----------
    results:
        Full list returned by matcher.run.
    df_a, df_b:
        Original source DataFrames (needed to pull extra column values).
    extra_cols_a, extra_cols_b:
        User-selected extra columns to include (may be empty lists).
    output_path:
        Destination file path (.xlsx).
    active_tiers:
        Which tiers to include in the Matched sheet.  Defaults to
        [TIER_EXACT] so the sheet opens filtered to exact matches only.
        The Unmatched sheet always shows all No Match rows regardless.
    deduplicate:
        If True, resolve File B duplicates before writing: when multiple
        File A rows share the same best File B match, only the highest-
        scoring pair is kept; the others are downgraded to No Match.
    """
    if active_tiers is None:
        active_tiers = [TIER_EXACT]

    if deduplicate:
        results = deduplicate_results(results)

    wb = Workbook()
    # Remove the default blank sheet
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    ws_matched = wb.create_sheet("Matched")
    ws_unmatched = wb.create_sheet("Unmatched")

    _write_summary_sheet(ws_summary, results, output_path)
    _write_matched_sheet(
        ws_matched, results, df_a, extra_cols_a, df_b, extra_cols_b, active_tiers
    )
    _write_unmatched_sheet(ws_unmatched, results, df_a, extra_cols_a)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, results: list[MatchResult], output_path: str) -> None:
    # Title
    title_cell = ws.cell(row=1, column=1, value="Fuzzy Account Matcher — Results Summary")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 22

    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=1, value="Generated:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=ts)
    ws.cell(row=3, column=1, value="Output file:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=str(Path(output_path).name))

    # Algorithm metadata
    ws.cell(row=5, column=1, value="Algorithm weights").font = Font(bold=True, color="1F4E79")
    ws.cell(row=6, column=1, value="Token / Jaccard")
    ws.cell(row=6, column=2, value="40%")
    ws.cell(row=7, column=1, value="Ratio (character)")
    ws.cell(row=7, column=2, value="35%")
    ws.cell(row=8, column=1, value="Partial ratio")
    ws.cell(row=8, column=2, value="25%")
    ws.cell(row=9, column=1, value="Match threshold")
    ws.cell(row=9, column=2, value="≥ 0.60")

    # Tier counts
    ws.cell(row=11, column=1, value="Match Quality Summary").font = Font(bold=True, color="1F4E79")
    headers = ["Tier", "Count", "Percentage"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=12, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN

    total = len(results)
    counts: dict[str, int] = {t: 0 for t in TIERS}
    for r in results:
        counts[r.tier] += 1

    for row_offset, tier in enumerate(TIERS, start=0):
        cnt = counts[tier]
        pct = (cnt / total * 100) if total else 0.0
        row = 13 + row_offset
        ws.cell(row=row, column=1, value=tier).fill = _TIER_FILLS[tier]
        ws.cell(row=row, column=2, value=cnt)
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")

    # Total row
    total_row = 13 + len(TIERS)
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value="100.0%").font = Font(bold=True)

    _auto_fit_columns(ws)


def _write_matched_sheet(
    ws,
    results: list[MatchResult],
    df_a: pd.DataFrame,
    extra_cols_a: list[str],
    df_b: pd.DataFrame,
    extra_cols_b: list[str],
    active_tiers: list[str],
) -> None:
    # Build header row
    fixed_headers = ["#", "File1_Name", "File2_Name", "Score", "Tier"]
    f1_headers = [f"F1_{c}" for c in extra_cols_a]
    f2_headers = [f"F2_{c}" for c in extra_cols_b]
    all_headers = fixed_headers + f1_headers + f2_headers
    n_cols = len(all_headers)

    for col, h in enumerate(all_headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"

    # Build index lookups for fast row retrieval
    df_a_indexed = df_a.set_index(df_a.index)
    df_b_indexed = df_b.set_index(df_b.index)

    # Write ALL tiers so the table contains every row; the AutoFilter
    # pre-selection tells Excel which tiers to show on open.
    for row_num, result in enumerate(results, start=2):
        fill = _TIER_FILLS[result.tier]
        idx = row_num - 1

        values = [
            idx,
            result.raw_a,
            result.raw_b,
            round(result.combined_score, 4),
            result.tier,
        ]

        # Extra columns from File A
        if extra_cols_a and result.index_a is not None:
            try:
                row_a = df_a_indexed.loc[result.index_a]
                values += [_safe_val(row_a[c]) for c in extra_cols_a]
            except (KeyError, IndexError):
                values += [""] * len(extra_cols_a)
        else:
            values += [""] * len(extra_cols_a)

        # Extra columns from File B
        if extra_cols_b and result.index_b is not None:
            try:
                row_b = df_b_indexed.loc[result.index_b]
                values += [_safe_val(row_b[c]) for c in extra_cols_b]
            except (KeyError, IndexError):
                values += [""] * len(extra_cols_b)
        else:
            values += [""] * len(extra_cols_b)

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = fill
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN

    _auto_fit_columns(ws)

    last_data_row = len(results) + 1  # +1 for header
    if last_data_row < 2:
        return  # empty — don't add a table

    table_ref = f"A1:{get_column_letter(n_cols)}{last_data_row}"

    # ── Hide rows whose tier is not in active_tiers ───────────────────────
    # Setting row_dimensions.hidden is the only reliable way to pre-apply a
    # filter so rows are already hidden when Excel opens the file — no
    # "Re-apply" click needed.  The Table's built-in autoFilter provides the
    # dropdown arrows; we do NOT set a duplicate worksheet-level auto_filter
    # or inject FilterColumn XML (both corrupt the file under Excel's strict
    # OOXML validation).
    active_set = set(active_tiers)
    for row_num, result in enumerate(results, start=2):
        if result.tier not in active_set:
            ws.row_dimensions[row_num].hidden = True

    # ── Excel Table ────────────────────────────────────────────────────────
    # Wrapping the range as a Table gives sortable/filterable column headers
    # and row striping.  The Table element owns the autoFilter; we let
    # openpyxl write the default (no filterColumn entries) so Excel doesn't
    # see conflicting filter specs.
    table = Table(displayName="MatchedResults", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _write_unmatched_sheet(
    ws,
    results: list[MatchResult],
    df_a: pd.DataFrame,
    extra_cols_a: list[str],
) -> None:
    fixed_headers = ["#", "File1_Name"]
    f1_headers = [f"F1_{c}" for c in extra_cols_a]
    all_headers = fixed_headers + f1_headers
    n_cols = len(all_headers)

    for col, h in enumerate(all_headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"

    unmatched = [r for r in results if r.tier == TIER_NO_MATCH]
    df_a_indexed = df_a.set_index(df_a.index)
    fill = _TIER_FILLS[TIER_NO_MATCH]

    for row_num, result in enumerate(unmatched, start=2):
        idx = row_num - 1
        values: list = [idx, result.raw_a]

        if extra_cols_a and result.index_a is not None:
            try:
                row_a = df_a_indexed.loc[result.index_a]
                values += [_safe_val(row_a[c]) for c in extra_cols_a]
            except (KeyError, IndexError):
                values += [""] * len(extra_cols_a)
        else:
            values += [""] * len(extra_cols_a)

        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = fill
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN

    _auto_fit_columns(ws)

    last_data_row = len(unmatched) + 1
    if last_data_row < 2:
        return

    table_ref = f"A1:{get_column_letter(n_cols)}{last_data_row}"
    table = Table(displayName="UnmatchedRows", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium3",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws) -> None:
    """Estimate column widths from cell content; cap at _MAX_COL_WIDTH."""
    col_widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            length = len(str(cell.value))
            current = col_widths.get(col_letter, _MIN_COL_WIDTH)
            col_widths[col_letter] = min(_MAX_COL_WIDTH, max(current, length + 2))

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def _safe_val(value):
    """Convert a pandas cell value to a Python scalar safe for openpyxl."""
    if pd.isna(value) if not isinstance(value, (list, dict)) else False:
        return ""
    if hasattr(value, "item"):  # numpy scalar
        return value.item()
    return value
