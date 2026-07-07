"""
Integration tests for core/exporter.py
"""
# ruff: noqa: E501
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from core.exporter import build_workbook
from models.match_result import (
    MatchResult,
    TIER_EXACT,
    TIER_GOOD,
    TIER_POSSIBLE,
    TIER_NO_MATCH,
    TIER_FILL_COLOURS,
)


def _make_results() -> list[MatchResult]:
    tiers = [TIER_EXACT, TIER_GOOD, TIER_POSSIBLE, TIER_NO_MATCH]
    results = []
    for i, tier in enumerate(tiers * 3):  # 12 results, 3 of each tier
        score_val = {TIER_EXACT: 0.98, TIER_GOOD: 0.85, TIER_POSSIBLE: 0.70, TIER_NO_MATCH: 0.40}[tier]
        results.append(
            MatchResult(
                index_a=i,
                index_b=i if tier != TIER_NO_MATCH else None,
                raw_a=f"Company A{i}",
                raw_b=f"Company B{i}" if tier != TIER_NO_MATCH else "",
                norm_a=f"company a{i}",
                norm_b=f"company b{i}",
                token_score=score_val,
                ratio_score=score_val,
                partial_score=score_val,
                combined_score=score_val,
                tier=tier,
            )
        )
    return results


def _make_dfs(results: list[MatchResult]):
    df_a = pd.DataFrame(
        {"Name": [r.raw_a for r in results], "Country": ["DE"] * len(results), "Revenue": range(len(results))},
        index=range(len(results)),
    )
    df_b = pd.DataFrame(
        {"Account": [r.raw_b for r in results], "Region": ["EU"] * len(results)},
        index=range(len(results)),
    )
    return df_a, df_b


class TestBuildWorkbook:
    def test_sheet_names(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, [], df_b, [], out)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "Matched" in wb.sheetnames
        assert "Unmatched" in wb.sheetnames

    def test_matched_sheet_row_count(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, [], df_b, [], out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Matched"]
        # header row + one row per result
        assert ws.max_row == len(results) + 1

    def test_unmatched_sheet_row_count(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, [], df_b, [], out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Unmatched"]
        no_match_count = sum(1 for r in results if r.tier == TIER_NO_MATCH)
        assert ws.max_row == no_match_count + 1  # +1 for header

    def test_extra_columns_in_matched_sheet(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, ["Country"], df_b, ["Region"], out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Matched"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "F1_Country" in headers
        assert "F2_Region" in headers

    def test_tier_cell_colours(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, [], df_b, [], out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Matched"]
        # Check the Tier column (column 5) for the first data row
        cell = ws.cell(2, 5)
        fill_colour = cell.fill.fgColor.rgb
        expected = "FF" + TIER_FILL_COLOURS[results[0].tier]  # openpyxl prepends alpha
        assert fill_colour == expected

    def test_deduplication_keeps_best_match(self, tmp_path):
        """When dedup=True, a File B row claimed by two File A rows → lower scorer loses."""
        from core.exporter import deduplicate_results
        from models.match_result import TIER_EXACT, TIER_GOOD, TIER_NO_MATCH

        r_winner = MatchResult(index_a=0, index_b=5, raw_a="Acme", raw_b="Acme", combined_score=0.99, tier=TIER_EXACT)
        r_loser  = MatchResult(index_a=1, index_b=5, raw_a="Acme2", raw_b="Acme", combined_score=0.80, tier=TIER_GOOD)
        r_unique = MatchResult(index_a=2, index_b=6, raw_a="Beta",  raw_b="Beta",  combined_score=0.95, tier=TIER_EXACT)

        deduped = deduplicate_results([r_winner, r_loser, r_unique])

        # winner keeps its match
        assert deduped[0].index_b == 5
        assert deduped[0].tier == TIER_EXACT
        # loser is downgraded
        assert deduped[1].index_b is None
        assert deduped[1].tier == TIER_NO_MATCH
        # unrelated row unaffected
        assert deduped[2].index_b == 6

    def test_deduplication_does_not_mutate_originals(self, tmp_path):
        from core.exporter import deduplicate_results
        from models.match_result import TIER_GOOD

        r = MatchResult(index_a=0, index_b=5, raw_a="X", raw_b="X", combined_score=0.80, tier=TIER_GOOD)
        deduplicate_results([r, r])  # same object twice — loser would be downgraded
        # original must be untouched
        assert r.index_b == 5
        assert r.tier == TIER_GOOD

    def test_summary_sheet_has_all_tiers(self, tmp_path):
        results = _make_results()
        df_a, df_b = _make_dfs(results)
        out = str(tmp_path / "out.xlsx")
        build_workbook(results, df_a, [], df_b, [], out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        # Collect all cell values in the sheet
        all_values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        for tier in (TIER_EXACT, TIER_GOOD, TIER_POSSIBLE, TIER_NO_MATCH):
            assert tier in all_values
